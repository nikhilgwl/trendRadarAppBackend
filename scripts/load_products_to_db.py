"""
One-time script: loads Nykaa_Pivoted_Wide.xlsx into the hul_products Supabase table.
Safe to re-run — it truncates and reloads.
Run from the backend/ directory: python scripts/load_products_to_db.py
"""
import os
import sys
import logging

# Allow imports from backend root
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from dotenv import load_dotenv

load_dotenv()
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

EXCEL_PATH = os.path.join(
    os.path.dirname(__file__), '..', '..', 'Nykaa_Pivoted_Wide.xlsx'
)
BATCH_SIZE = 100


def clean(val):
    """Return None for empty/nan strings, else strip the value."""
    if val is None:
        return None
    s = str(val).strip()
    return None if s in ('', 'nan', 'None', 'NaN') else s


def load_products():
    import db
    client = db.get_client()
    if not client:
        logger.error("No Supabase client — check SUPABASE_URL and SUPABASE_KEY in .env")
        return

    logger.info(f"Opening Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb['Nykaa_Wide']

    # Build column-name → index map from header row
    header_row = next(ws.iter_rows(values_only=True))
    col_idx = {name: i for i, name in enumerate(header_row) if name is not None}
    logger.info(f"Columns found: {sorted(col_idx.keys())}")

    def get(row, col_name):
        idx = col_idx.get(col_name)
        if idx is None or idx >= len(row):
            return None
        return clean(row[idx])

    # Truncate existing data before reload
    logger.info("Truncating hul_products table...")
    try:
        client.table("hul_products").delete().neq("id", 0).execute()
    except Exception as e:
        logger.warning(f"Truncate failed (may be empty): {e}")

    products = []
    skipped = 0
    total = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # skip header
            continue

        brand = get(row, 'brand')
        product_name = get(row, 'product_name')

        # Skip rows without essential fields
        if not brand or not product_name:
            skipped += 1
            continue

        # Gather keywords (keyword_1 through keyword_10)
        keywords = [
            get(row, f'keyword_{n}')
            for n in range(1, 11)
        ]
        keywords = [k for k in keywords if k]

        # Gather feature bullets
        bullets = [
            get(row, f'feature_bullet_{n}')
            for n in range(1, 7)
        ]
        bullets = [b for b in bullets if b]

        basepack_raw = row[col_idx.get('Basepack', 0)] if 'Basepack' in col_idx else None
        try:
            basepack = int(basepack_raw) if basepack_raw is not None else None
        except (ValueError, TypeError):
            basepack = None

        mrp_raw = get(row, 'mrp')
        try:
            mrp = float(mrp_raw) if mrp_raw else None
        except (ValueError, TypeError):
            mrp = None

        products.append({
            "basepack":            basepack,
            "brand":               brand,
            "product_name":        product_name,
            "segment":             get(row, 'segment'),
            "concern":             get(row, 'concern'),
            "key_ingredients":     get(row, 'key_ingredients'),
            "keywords":            keywords,
            "feature_bullets":     bullets,
            "product_description": get(row, 'product_description'),
            "skin_type":           get(row, 'skin_type'),
            "hair_type":           get(row, 'hair_type'),
            "formulation":         get(row, 'formulation'),
            "mrp":                 mrp,
        })
        total += 1

        # Batch insert
        if len(products) >= BATCH_SIZE:
            _insert_batch(client, products)
            products = []
            logger.info(f"  Inserted {total} products so far...")

    # Final batch
    if products:
        _insert_batch(client, products)

    wb.close()
    logger.info(f"Done. Inserted {total} products, skipped {skipped} empty rows.")


def _insert_batch(client, batch: list):
    try:
        client.table("hul_products").insert(batch).execute()
    except Exception as e:
        logger.error(f"Batch insert failed: {e}")
        # Try row-by-row to isolate bad rows
        for row in batch:
            try:
                client.table("hul_products").insert(row).execute()
            except Exception as row_err:
                logger.warning(f"  Skipping row {row.get('product_name')}: {row_err}")


if __name__ == "__main__":
    load_products()
